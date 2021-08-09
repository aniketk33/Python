from translate import Translator
from openpyxl import load_workbook

workbook_name = 'MultiLang.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

class clsTranslate():
    def translateText(self, strString, strTolang):
        self.strString = strString
        self.strTolang = strTolang
        translator = Translator(to_lang=self.strTolang)
        translation = translator.translate(self.strString)
        return (str(translation))


textToBeTranslated = []
for i in range(2,page.max_row+1):
    data = page.cell(row=i,column=1)
    if(data.value is None):
        continue    
    textToBeTranslated.append(data.value)

objTrans=clsTranslate()
lang=[]
if(len(textToBeTranslated) != 0):
    #assign multi lang data to lang list
    for data in textToBeTranslated:
        arabicText = objTrans.translateText(data,'ar')
        chineseText = objTrans.translateText(data,'zh-hans')
        frenchText = objTrans.translateText(data,'fr')
        russianText = objTrans.translateText(data,'ru')
        spanishText = objTrans.translateText(data,'es')
        lang.append(arabicText)
        lang.append(chineseText)
        lang.append(frenchText)
        lang.append(russianText)
        lang.append(spanishText)

if(len(lang) != 0):
    i = 0
    #assign multi lang values to particular row in the excel sheet
    for row in range(2,page.max_row+1):
        for col in range(2,page.max_column+1):
            page.cell(row=row,column=col).value = lang[i]
            i=i+1

wb.save(filename = workbook_name)