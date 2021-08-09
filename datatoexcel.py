from os import path
from googletrans import Translator, constants
from openpyxl import load_workbook

translator = Translator()
workbook_name = 'MultiLang.xlsx'
wb = load_workbook(workbook_name)
page = wb.active

textToBeTranslated = []
for i in range(2,page.max_row+1):
    data = page.cell(row=i,column=1)
    if(data.value is None):
        continue    
    textToBeTranslated.append(data.value)

lang=[]
if(len(textToBeTranslated) != 0):
    #assign multi lang data to lang list
    for data in textToBeTranslated:
        arabicText = translator.translate(data,dest='ar')
        chineseText = translator.translate(data,dest='zh-cn')
        frenchText = translator.translate(data,dest='fr')
        russianText = translator.translate(data,dest='ru')
        spanishText = translator.translate(data,dest='es')
        lang.append(arabicText.text)
        lang.append(chineseText.text)
        lang.append(frenchText.text)
        lang.append(russianText.text)
        lang.append(spanishText.text)

if(len(lang) != 0):
    i = 0
    #assign multi lang values to particular row in the excel sheet
    for row in range(2,page.max_row+1):
        for col in range(2,page.max_column+1):
            page.cell(row=row,column=col).value = lang[i]
            i=i+1

wb.save(filename = workbook_name)